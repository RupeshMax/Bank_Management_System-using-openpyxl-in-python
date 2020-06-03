from openpyxl import load_workbook

class Account:
    accno=0
    name=''
    phone=0
    mode=''
    bal=0

    def create_account(self,sno):
        self.accno=sno+1
        self.name=input("Applicant Name:")
        self.phone=int(input("Phone Number:"))
        self.mode=input("saving or current account(s/c):")
        self.bal=int(input("Amount (minimum=1000):"))
        print("Account Created Successfully!...\n***Your Account Number is",self.accno,"***")

    def storing(self):
        sheet.append((self.accno,self.name,self.phone,self.mode,self.bal))
        

def Close_account(accno):
    max_row=sheet.max_row
    flag=0
    for i in range(2,max_row+1):
        if sheet.cell(row=i,column=1).value==accno:
            sheet.delete_rows(idx=i)
            flag=1
            print("Your Account successfully closed!...")
    if flag==0:
        print("please check your Account Number!...")
        if input("do you want to continue ?(y/n)")=='y':
            accno=int(input("Enter the correct Account Number : "))
            Close_account(accno)



def checkbalance(accno):
    max_row=sheet.max_row
    max_column=sheet.max_column
    flag=0
    for i in range(2,max_row+1):
        if sheet.cell(row=i,column=1).value==accno:
            flag=1
            
            print("\nYour Main Balance : ",sheet.cell(row=i,column=max_column).value)
    if flag==0:
        print("please check your Account Number!..\ndo you want to continue(y/n)")
        if input()=='y':
            accno=int(input("Enter the correct Account Number"))
            checkbalance(accno)
    
def withdraw(accno):
    max_row=sheet.max_row
    max_column=sheet.max_column
    flag=0
    for i in range(2,max_row+1):
        if sheet.cell(row=i,column=1).value==accno:
            flag=1
            balance=sheet[str(chr(64+max_column))+str(i)].value
            amount=int(input("Enter the withdraw amount :"))
            if amount < balance:
                sheet[str(chr(64+max_column))+str(i)]=balance-amount
            else:
                print("Insufficient balance!..")
                
    if flag==0:
        print("please check your Account Number!..\ndo you want to continue(y/n)")
        if input()=='y':
            accno=int(input("Enter the correct Account Number"))
            withdraw(accno)

def deposit(accno):
    max_row=sheet.max_row
    max_column=sheet.max_column
    flag=0
    for i in range(2,max_row+1):
        if sheet.cell(row=i,column=1).value==accno:
            flag=1
            balance=sheet[str(chr(64+max_column))+str(i)].value
            amount=int(input("Enter the deposit amount :"))
            if amount < balance:
                sheet[str(chr(64+max_column))+str(i)]=balance+amount
            
    if flag==0:
        print("please check your Account Number!..\ndo you want to continue(y/n)")
        if input()=='y':
            accno=int(input("Enter the correct Account Number"))
            deposit(accno)
    
    
            
def displaying(sheet):
    print("\n")
    max_row=sheet.max_row
    max_column=sheet.max_column
    for i in range(1,max_row+1):
        for j in range(1,max_column+1):
            cell_obj=sheet.cell(row=i,column=j)
            print("{0:14}".format(cell_obj.value),end=' | ')
        print()
    print("\n")


    
    
wb=load_workbook(filename='database.xlsx')
sheet=wb.active
print("\tBank Management System")

while True:
    print("----"*10+"\nchoose your option :")
    print("\t1. Create Account\n\t2. Check Balance\n\t3. Deposit\n\t4. Withdraw\n\t5. Modify an Account\n\t6. close an Account\n\t7. Admin Login\n\t8. Exit")
    x=int(input("Enter your choosen number: "))
    print("----"*10+"\n")
    if x==1:
        account=Account()
        max_row=sheet.max_row
        accno=0
        for i in range(2,max_row+1):
            if sheet.cell(row=i,column=1).value>accno:
                accno=sheet.cell(row=i,column=1).value
        account.create_account(accno)
        account.storing()
        wb.save('database.xlsx')

    elif x==2:
        accno=int(input("Enter the your Account Number : "))
        checkbalance(accno)

    elif x==3:
        accno=int(input("Enter the your Account Number : "))
        deposit(accno)
        
    elif x==4:
        accno=int(input("Enter the your Account Number : "))
        withdraw(accno)
        
    elif x==6:
        accno=int(input("Enter the your Account Number : "))
        Close_account(accno)
        wb.save('database.xlsx')
        
    elif x==7:
        print("\nchoose your option :")
        print("\t1. Customer database\n\t2. Check Balance")
        y=int(input("Enter your choosen number: "))
        if y==1:
            displaying(sheet)
        
    elif x==8:
        print("Thank you")
        break
        
